import os
import time
from pytube import Playlist, YouTube
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import requests
from io import BytesIO


def sanitize_filename(filename):
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    return filename


def get_playlist_title(url):
    try:
        playlist = Playlist(url)
        return playlist.title
    except Exception as e:
        print(f"Error retrieving playlist title: {e}")
        return None


def get_excel_path(directory, playlist_title):
    sanitized_title = sanitize_filename(playlist_title)
    filename = f"{sanitized_title}.xlsx"
    return os.path.join(directory, filename)


def download_thumbnail(url):
    response = requests.get(url)
    if response.status_code == 200:
        return BytesIO(response.content)
    else:
        return None


def add_video_to_excel(workbook, video, row, excel_path):
    ws = workbook.active
    thumbnail_io = download_thumbnail(video['thumbnail_url'])
    if thumbnail_io:
        img = Image(thumbnail_io)
        ws.add_image(img, f'A{row}')
        ws.row_dimensions[row].height = 90  # Adjust row height for the image
        ws.column_dimensions[get_column_letter(
            1)].width = 20  # Adjust column width

    ws[f'B{row}'] = video['title']
    ws[f'C{row}'] = video['url']
    # Save after each video to update the file progressively
    workbook.save(filename=excel_path)


def main():
    playlist_url = input("Enter the YouTube playlist URL: ")
    directory = input(
        "Enter the directory where you want to save the Excel file: ")

    os.makedirs(directory, exist_ok=True)

    playlist_title = get_playlist_title(playlist_url)
    if playlist_title is None:
        print("Failed to retrieve playlist title. Exiting.")
        return

    excel_path = get_excel_path(directory, playlist_title)
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Thumbnail", "Title", "URL"])

    row = 2  # Start from the second row, after the header
    for url in Playlist(playlist_url).video_urls:
        try:
            time.sleep(2)  # Delay to avoid rate limits
            video = YouTube(url)
            print(f"Retrieving: {video.title}")
            new_video = {
                'title': video.title,
                'url': video.watch_url,
                'thumbnail_url': video.thumbnail_url
            }
            add_video_to_excel(workbook, new_video, row, excel_path)
            row += 1
        except Exception as e:
            print(f"Error retrieving video details: {e}")


if __name__ == "__main__":
    main()
